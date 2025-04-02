import fitz  # PyMuPDF
import re
import os
from openpyxl import load_workbook

def my_extraction_process_barclays(n, doc, pdf_name, sheet):
    page = doc[n]
    my_text = page.get_text().strip()  
    my_text = re.sub(r'(\n+)', ' * ', my_text)  

    first_pattern = r'(Terms\s+used\s+.*?\s+supplement\.)'  
    matched = re.findall(first_pattern, my_text, re.IGNORECASE)

    if matched:
        my_patterns = [  
            r'\*\s+[iI]ssuer\:\s+\*(.*?)\*',  
            r'\*\s+[dD]enominations\:\s+\*(.*?)\*',  
            r'\*\s+[iI]nitial\s+[vV]aluation\s+[dD]ate\:\s+\*(.*?)\*',  
            r'\*\s+[iI]ssue\s+[dD]ate\:\s+\*(.*?)\*',  
            r'\*\s+[fF]inal\s+[vV]aluation\s+[dD]ate\:\*\s+\*(.*?)\*',  
            r'\*\s+[mM]aturity\s+[dD]ate\:\*\s+\*(.*?)\*',  
            r'\*\s+[rR]eference\s+[aA]ssets\:\s+\*(.*?)\*\s+[pP]ayment',  
        ]
        
        my_values = []
        
        for pattern in my_patterns:
            match = re.findall(pattern, my_text)
            my_values.append(match[0].strip() if match else "N/A")  # Default to "N/A" if no match
        
        # Headers in Column A & Extracted Values in Column B
        headers = [
            "PDF name", "Page no.", "Issuer", "Denominations", 
            "Initial Valuation Date", "Issue Date", "Final Valuation Date", 
            "Maturity Date", "Reference Assets"
        ]
        values = [pdf_name, n + 1] + my_values  # Include PDF name and page number

        # Write headers in Column A and corresponding values in Column B
        start_row = sheet.max_row + 2  # Find next available row to start
        for i, header in enumerate(headers):
            sheet.cell(row=start_row + i, column=1, value=header)  # Column A
            sheet.cell(row=start_row + i, column=2, value=values[i])  # Column B

def process_pdfs(directory, excel_path):
    for root, dirs, files in os.walk(directory):
        for filename in files:
            if filename.endswith(".pdf"):  
                pdf_path = os.path.join(root, filename)

                doc = fitz.open(pdf_path)

                # Load Excel file
                workbook = load_workbook(filename=excel_path)
                sheet = workbook["Barclays"]

                print(f"Processing: {filename}")

                for i in range(len(doc)):
                    my_extraction_process_barclays(i, doc, filename, sheet)

                workbook.save(excel_path)
                print(f"Updated: {filename}")

# File paths
directory = "C:\Barclays_Sheet\Barclays Termsheet"
excel_path = "C:/Barclays_Sheet/Trial.xlsx"

# Run extraction
process_pdfs(directory, excel_path)
