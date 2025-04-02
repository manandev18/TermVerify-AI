import fitz
import re
import os
from openpyxl import load_workbook

def my_extraction_process_barclays(n,doc,pdf_name):
    page = doc[n]
    
    my_text = page.get_text()
    my_text = my_text.strip() 
    my_text = re.sub(r'(\n+)', ' * ', my_text) 
    first_pattern = r'(Terms\s+used\s+.*?\s+supplement\.)' 
    matched = re.findall(first_pattern, my_text, re.IGNORECASE)
    if matched:
        my_patterns = [ r'\*\s+[iI]ssuer\:\s+\*(.*?)\*',  
                        r'\*\s+[dD]enominations\:\s+\*(.*?)\*',  
                        r'\*\s+[iI]nitial\s+[vV]aluation\s+[dD]ate\:\s+\*(.*?)\*',   
                        r'\*\s+[iI]ssue\s+[dD]ate\:\s+\*(.*?)\*',  
                        r'\*\s+[fF]inal\s+[vV]aluation\s+[dD]ate\:\*\s+\*(.*?)\*',  
                        r'\*\s+[mM]aturity\s+[dD]ate\:\*\s+\*(.*?)\*',  
                        r'\*\s+[rR]eference\s+[aA]ssets\:\s+\*(.*?)\*\s+[pP]ayment',   
                        ]
        my_outcomes = [0]*len(my_patterns)
        my_values1 = [0]*len(my_patterns)
        for i in range(0,len(my_patterns)):
            pattern = my_patterns[i]
            matches = re.findall(pattern, my_text)
            if matches:
                if (matches[0] != None) and (type(matches[0]) != str):
                    matches[0] = str(matches[0])
                try:
                    if matches != None:
                        my_outcomes[i] = matches[0]
                except:
                    print("Error 1 at element - " + str(i))
        
        for i in range(0,len(my_outcomes)):
            if type(my_outcomes[i]) != str:
                my_outcomes[i] = str(my_outcomes[i])
            my_values1[i] = re.sub(r'\s\*', '', my_outcomes[i])

        workbook = load_workbook(filename="C:/Barclays_Sheet/Trial.xlsx")
        sheet = workbook["Barclays"]
                    
        if my_values1[6] != 0:
            my_values1.insert(0," ")
            my_values1.insert(1,pdf_name)
            my_values1.insert(2,n+1)
            sheet.append(my_values1)
            workbook.save("C:/Barclays_Sheet/Trial.xlsx")
    else:
        matched = None
        
directory = "C:/Barclays/Barclays Termsheet"

for root, dirs, files in os.walk(directory):
    for filename in files:
        print("directory accessed")
        pdf_path = os.path.join(root, filename)

        doc = fitz.open(pdf_path)
        workbook = load_workbook(filename="C:/Barclays_Sheet/Trial.xlsx")
        sheet = workbook["Barclays"]
        file_name = os.path.basename(pdf_path)
        pdf_data = [[" "],[" "],[" "],[" ","PDF name","page no.","Issuer","Denominations","Initial Valuation Date","Issue Date","Final Valuation Date","Maturity Date","Reference Assets"]]
        for data in pdf_data:
            sheet.append(data)
        workbook.save("C:/Barclays_Sheet/Trial.xlsx")
        pattern = '[0-9].*[0-9]'
        matching = re.findall(pattern,filename)
        
        if matching[0]:
            print(f"Processing {matching[0]} file: {file_name}")
            for i in range(len(doc)):
                my_extraction_process_barclays(i,doc,file_name)
            print(f"Updated sheet for {matching[0]} file: {file_name}")
